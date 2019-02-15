from openpyxl import load_workbook

def copyData(sourceFilePath,destinationFilePath):
    wb = load_workbook(sourceFilePath)
    sheet = wb.sheetnames
    dataSheet = ''.join(sheet)
    sheet = wb[dataSheet]
	# get max row count
    max_row = sheet.max_row
    #get max column count
    max_column = sheet.max_column
    a = ''
    for i in range(1, max_row + 1):
        for j in range(1, max_column + 1):
            cell_obj = sheet.cell(row=i, column=j)
            result = str(cell_obj.value)
            a = a + result + '|'
        a = a + '\n'
        with open(destinationFilePath,"a") as myfile:
            myfile.write(a)
        a = ''
def main():
    copyData("..//SpreadSheet//SampleData.xlsx","..//SpreadSheet//ResultData.txt")

if __name__ == '__main__':
    main()
		



